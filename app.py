import os
import tkinter as tk
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
from datetime import datetime, time
from tkinter import *
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

#---------========-----------=====--------##---------========-----------=====--------##---------========-----------=====--------#

arquivoExcel = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'temperaturas.xlsx')  # Arquivo Excel onde serão salvos os dados

root = tk.Tk()
pointer = 1


root.title("Temperature Catcher - São Paulo")
#root.configure(background="white")
root.minsize(200, 200)
root.maxsize(500, 500)
root.geometry("100x300+25+25")

tk.Label(root, text = "Atualizar Previsão na planilha")

curTime =datetime.now() # Salvando data e hora na variavel curTime
curTimeFormated = curTime.strftime("%Y-%m-%d %H:%M:%S")



def buttonPressed():
    temperatura, qtdQueimadas = obter_temperatura_sao_paulo()
    salvar_temperatura_em_excel(arquivoExcel, temperatura, qtdQueimadas)
    


# Função responsavel por retornar a temperatura no site CLima Tempo. utilização do selenium
def obter_temperatura_sao_paulo():  
    options = Options()
    options.add_argument("--headless") ## COM ESSE ARGUMENTO, PODEMOS RODAR SEM ABRIR O NAVEGADOR. OPCIONAL 
    options.add_argument("--disable-gpu") ## COM ESSE ARGUMENTO, PODEMOS RODAR SEM ABRIR O NAVEGADOR. OPCIONAL 
    options.add_argument("--no-sandbox") ## COM ESSE ARGUMENTO, PODEMOS RODAR SEM ABRIR O NAVEGADOR. OPCIONAL 

    driver = webdriver.Chrome(options=options) # AQUI DEFINIMOS AS OPÇ˜OES DEFINIDAS ACIMA DENTRO DO WEBDRIVER DO CHROME.
    try:
        driver.get("https://www.google.com/search?q=previsao+do+tempo+em+sao+paulo") 
        driver.get("https://www.climatempo.com.br/previsao-do-tempo/agora/cidade/558/saopaulo-sp")

        # aqui aguardo 5 segundos o carregamento da pagina e busco o elemento pelo seletor CSS, com as classes da tag span
        temperatura = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "span.-bold.-gray-dark-2.-font-55._margin-l-20._center"))).text
        
        queimadas = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "recentFire"))).text
        
        #retorno do valor da temperatura para usar n a função intermediaria (que chama as duas.)
        return temperatura, queimadas
    finally:
        driver.quit()



def salvar_temperatura_em_excel(caminho_arquivo, temperatura, queimadas):
    # Abre a planilha existente
    planilhaCaminho = load_workbook(caminho_arquivo)
    planilhaAba = planilhaCaminho.active  # Usa a primeira aba (sheet)

    # Encontra a próxima linha vazia
    proxima_linha = planilhaAba.max_row + 1

    # Insere a data/hora e a temperatura
    planilhaAba.cell(row=proxima_linha, column=1, value=curTime) # CurTime esta sendo definida acima, e mostra a data e hora no bottom da interface
    planilhaAba.cell(row=proxima_linha, column=2, value=temperatura)
    planilhaAba.cell(row=proxima_linha, column=3, value=queimadas)

    # Salva o arquivo e exibe no terminal
    planilhaCaminho.save(caminho_arquivo)
    print(f"Temperatura {temperatura}°C salva na linha {proxima_linha}.") # Aqui confirmamos no terminal se a temperatura foi salva, e em qual linha.


tk.Label(root, text=curTimeFormated).pack(side=tk.BOTTOM) #Salvando data e hora no final da janela


tk.Label(root, text='Buscar dados de Temperatura').pack()
searchButton=tk.Button(root,text="Buscar", command=buttonPressed).pack()


# Funcão intermediaria. Crirei para que ao apertar o botão, ambas as funções sejam chamadas com seus respectivos parametos.

root.mainloop()



